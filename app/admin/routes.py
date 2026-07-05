import os
import io
import csv
import json
import zipfile
import tempfile
from datetime import datetime, timedelta
from flask import (Blueprint, render_template, request, flash, redirect,
                   url_for, abort, jsonify, Response, send_file, current_app)
from flask_login import login_required, current_user
from ..extensions import db
from ..models.models import (User, Attraction, Guide, Hotel, Restaurant,
                              Event, News, Booking, Review, AuditLog, QRCode,
                              CMSSetting, Category, Notification, FAQ,
                              Partner, Advertisement, Payment, HotelRoom,
                              RestaurantMenu, NewsComment)
from ..utils.decorators import admin_required
from ..utils.helpers import log_action, save_uploaded_file, delete_uploaded_file, allowed_file

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def _require_admin():
    if not current_user.is_authenticated or not current_user.is_admin:
        abort(403)


# ─── DASHBOARD ─────────────────────────────────────────────────────────────

@admin_bp.route("/")
@login_required
@admin_required
def dashboard():
    stats = {
        "users": User.query.filter(User.status != "deleted").count(),
        "attractions": Attraction.query.filter_by(status="active").count(),
        "guides": Guide.query.filter_by(verification_status="Approved", status="active").count(),
        "hotels": Hotel.query.filter_by(status="active").count(),
        "restaurants": Restaurant.query.filter_by(status="active").count(),
        "events": Event.query.filter_by(status="active").count(),
        "bookings": Booking.query.filter_by(status="active").count(),
        "pending_bookings": Booking.query.filter_by(booking_status="Pending", status="active").count(),
        "reviews": Review.query.filter_by(status="active").count(),
        "news": News.query.filter_by(status="active").count(),
        "qr_scans": db.session.query(db.func.sum(QRCode.scanned_count)).scalar() or 0,
        "revenue": db.session.query(db.func.sum(Payment.amount)).scalar() or 0,
    }
    recent_bookings = Booking.query.filter_by(status="active").order_by(Booking.created_at.desc()).limit(5).all()
    recent_users = User.query.filter(User.status != "deleted").order_by(User.created_at.desc()).limit(5).all()
    recent_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(8).all()
    top_attractions = Attraction.query.filter_by(status="active").order_by(Attraction.views.desc()).limit(5).all()

    # Monthly booking trend (last 6 months)
    trend_data = []
    for i in range(5, -1, -1):
        month_start = (datetime.utcnow().replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1)
        count = Booking.query.filter(Booking.created_at >= month_start, Booking.created_at < month_end).count()
        trend_data.append({"month": month_start.strftime("%b"), "count": count})

    return render_template(
        "admin/dashboard.html",
        stats=stats,
        recent_bookings=recent_bookings,
        recent_users=recent_users,
        recent_logs=recent_logs,
        top_attractions=top_attractions,
        trend_data=json.dumps(trend_data),
    )


# ─── USERS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/users")
@login_required
@admin_required
def users():
    page = request.args.get("page", 1, type=int)
    search_q = request.args.get("q", "").strip()
    role_filter = request.args.get("role", "")
    query = User.query.filter(User.status != "deleted")
    if search_q:
        query = query.filter((User.email.ilike(f"%{search_q}%")) | (User.first_name.ilike(f"%{search_q}%")) | (User.last_name.ilike(f"%{search_q}%")))
    if role_filter:
        query = query.filter_by(role=role_filter)
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template("admin/users.html", users=pagination.items, pagination=pagination, search_q=search_q, role_filter=role_filter)


@admin_bp.route("/users/new", methods=["GET", "POST"])
@login_required
@admin_required
def user_create():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        if User.query.filter_by(email=email).first():
            flash("Email already registered.", "danger")
            return redirect(url_for("admin.user_create"))
        u = User(
            email=email,
            first_name=request.form.get("first_name", "").strip(),
            last_name=request.form.get("last_name", "").strip(),
            role=request.form.get("role", "Tourist"),
            phone=request.form.get("phone", "").strip(),
            is_verified=request.form.get("is_verified") == "on",
            status="active",
        )
        u.set_password(request.form.get("password", "changeme123"))
        db.session.add(u)
        db.session.commit()
        log_action(f"Admin created user: {email}", module="admin.users", record_id=u.id)
        flash(f"User {u.full_name} created.", "success")
        return redirect(url_for("admin.users"))
    return render_template("admin/user_form.html", user=None, title="Create User")


@admin_bp.route("/users/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def user_edit(id):
    u = User.query.get_or_404(id)
    if request.method == "POST":
        u.first_name = request.form.get("first_name", u.first_name).strip()
        u.last_name = request.form.get("last_name", u.last_name).strip()
        u.role = request.form.get("role", u.role)
        u.phone = request.form.get("phone", u.phone or "").strip()
        u.is_verified = request.form.get("is_verified") == "on"
        u.heritage_points = int(request.form.get("heritage_points", u.heritage_points))
        new_pass = request.form.get("password", "").strip()
        if new_pass:
            u.set_password(new_pass)
        db.session.commit()
        log_action(f"Admin edited user #{id}", module="admin.users", record_id=id)
        flash("User updated.", "success")
        return redirect(url_for("admin.users"))
    return render_template("admin/user_form.html", user=u, title="Edit User")


@admin_bp.route("/users/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def user_delete(id):
    u = User.query.get_or_404(id)
    if u.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin.users"))
    u.soft_delete()
    db.session.commit()
    log_action(f"Admin soft-deleted user #{id} ({u.email})", module="admin.users", record_id=id)
    flash(f"User {u.email} deactivated.", "info")
    return redirect(url_for("admin.users"))


# ─── ATTRACTIONS ────────────────────────────────────────────────────────────

@admin_bp.route("/attractions")
@login_required
@admin_required
def attractions():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    query = Attraction.query.filter_by(status="active")
    if q:
        query = query.filter(Attraction.name.ilike(f"%{q}%"))
    pagination = query.order_by(Attraction.created_at.desc()).paginate(page=page, per_page=15, error_out=False)
    return render_template("admin/attractions.html", attractions=pagination.items, pagination=pagination, q=q)


@admin_bp.route("/attractions/new", methods=["GET", "POST"])
@login_required
@admin_required
def attraction_create():
    categories = Category.query.filter_by(type="attraction", status="active").all()
    if request.method == "POST":
        try:
            # Handle image upload - only from file upload
            image = request.files.get("image")
            image_url = None
            if image and allowed_file(image.filename):
                image_url = save_uploaded_file(image, "attractions")
            
            a = Attraction(
                name=request.form.get("name", "").strip(),
                category_id=request.form.get("category_id", type=int),
                description=request.form.get("description", ""),
                address=request.form.get("address", ""),
                history=request.form.get("history", ""),
                opening_hours=request.form.get("opening_hours", ""),
                ticket_price=float(request.form.get("ticket_price", 0)),
                image_url=image_url,
                latitude=float(request.form.get("latitude", 6.34)),
                longitude=float(request.form.get("longitude", 5.62)),
                contact=request.form.get("contact", ""),
                website=request.form.get("website", ""),
                featured=request.form.get("featured") == "on",
                active=request.form.get("active") == "on",
                status="active",
            )
            db.session.add(a)
            db.session.commit()
            
            # Generate QR code
            _generate_qr(a)
            
            log_action(f"Admin created attraction: {a.name}", module="admin.attractions", record_id=a.id)
            flash(f"Attraction '{a.name}' created with QR code.", "success")
            return redirect(url_for("admin.attractions"))
            
        except Exception as e:
            current_app.logger.error(f"Error creating attraction: {e}")
            flash(f"Error creating attraction: {str(e)}", "danger")
            db.session.rollback()
    
    return render_template("admin/attraction_form.html", attraction=None, categories=categories, title="New Attraction")


@admin_bp.route("/attractions/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def attraction_edit(id):
    a = Attraction.query.get_or_404(id)
    categories = Category.query.filter_by(type="attraction", status="active").all()
    
    if request.method == "POST":
        try:
            # Update basic fields
            a.name = request.form.get("name", a.name).strip()
            a.category_id = request.form.get("category_id", a.category_id, type=int)
            a.description = request.form.get("description", a.description)
            a.address = request.form.get("address", a.address)
            a.history = request.form.get("history", a.history)
            a.opening_hours = request.form.get("opening_hours", a.opening_hours)
            a.ticket_price = float(request.form.get("ticket_price", a.ticket_price or 0))
            a.latitude = float(request.form.get("latitude", a.latitude or 6.34))
            a.longitude = float(request.form.get("longitude", a.longitude or 5.62))
            a.contact = request.form.get("contact", a.contact)
            a.website = request.form.get("website", a.website)
            a.featured = request.form.get("featured") == "on"
            a.active = request.form.get("active") == "on"
            
            # Handle image upload - only from file upload
            image = request.files.get("image")
            if image and allowed_file(image.filename):
                # Delete old image if exists
                if a.image_url:
                    delete_uploaded_file(a.image_url, "attractions")
                # Save new image
                a.image_url = save_uploaded_file(image, "attractions")
            
            db.session.commit()
            
            # Regenerate QR code
            _generate_qr(a)
            
            log_action(f"Admin edited attraction #{id}", module="admin.attractions", record_id=id)
            flash("Attraction updated successfully.", "success")
            return redirect(url_for("admin.attractions"))
            
        except Exception as e:
            current_app.logger.error(f"Error updating attraction {id}: {e}")
            flash(f"Error updating attraction: {str(e)}", "danger")
            db.session.rollback()
    
    return render_template("admin/attraction_form.html", attraction=a, categories=categories, title="Edit Attraction")


@admin_bp.route("/attractions/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def attraction_delete(id):
    a = Attraction.query.get_or_404(id)
    # Delete image if exists
    if a.image_url:
        delete_uploaded_file(a.image_url, "attractions")
    a.soft_delete()
    db.session.commit()
    log_action(f"Admin deleted attraction #{id}", module="admin.attractions", record_id=id)
    flash("Attraction removed.", "info")
    return redirect(url_for("admin.attractions"))


@admin_bp.route("/attractions/<int:id>/qr-regenerate", methods=["POST"])
@login_required
@admin_required
def attraction_regen_qr(id):
    a = Attraction.query.get_or_404(id)
    try:
        filename = _generate_qr(a)
        if filename:
            flash(f"QR code regenerated successfully for {a.name}.", "success")
            log_action(f"Admin regenerated QR for attraction #{id}", module="admin.attractions", record_id=id)
        else:
            flash("Failed to regenerate QR code. Check logs for details.", "error")
    except Exception as e:
        current_app.logger.error(f"Error regenerating QR for attraction {id}: {e}")
        flash(f"Error regenerating QR code: {str(e)}", "error")
    
    return redirect(url_for("admin.attractions"))


def _generate_qr(attraction):
    """Generate QR code for an attraction and save to static/uploads/qr_codes/"""
    try:
        import qrcode
        from PIL import Image
        
        # Get the upload folder from config or use default
        upload_folder = current_app.config.get("UPLOAD_FOLDER", os.path.join(current_app.root_path, "static/uploads"))
        qr_dir = os.path.join(upload_folder, "qr_codes")
        os.makedirs(qr_dir, exist_ok=True)
        
        # Generate the URL for the attraction
        base_url = current_app.config.get("BASE_URL", "http://localhost:5000")
        url = f"{base_url}/attractions/{attraction.id}"
        
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(url)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        filename = f"qr_{attraction.id}.png"
        filepath = os.path.join(qr_dir, filename)
        img.save(filepath, 'PNG')
        
        # The path to store in database (relative to static folder)
        path = f"/static/uploads/qr_codes/{filename}"
        
        # Update or create QR code record
        existing = QRCode.query.filter_by(attraction_id=attraction.id).first()
        if existing:
            existing.code = filename
            existing.qr_image_path = path
            existing.url = url
        else:
            qr_record = QRCode(
                attraction_id=attraction.id,
                code=filename,
                qr_image_path=path,
                url=url,
                status="active"
            )
            db.session.add(qr_record)
        db.session.commit()
        
        current_app.logger.info(f"QR code generated for attraction {attraction.id}: {filename}")
        return filename
        
    except ImportError as e:
        current_app.logger.error(f"QRCode library not installed: {e}")
        flash("QRCode library not installed. Please install: pip install qrcode Pillow", "error")
        return None
    except Exception as e:
        current_app.logger.error(f"QR generation failed for attraction {attraction.id}: {e}")
        return None


# ─── GUIDES ─────────────────────────────────────────────────────────────────

@admin_bp.route("/guides")
@login_required
@admin_required
def guides():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "")
    query = Guide.query.filter_by(status="active")
    
    if q:
        query = query.join(User).filter(
            (User.first_name.ilike(f"%{q}%")) | 
            (User.last_name.ilike(f"%{q}%")) | 
            (User.email.ilike(f"%{q}%")) |
            (Guide.phone.ilike(f"%{q}%"))  # <-- Added phone search
        )
    if status_filter:
        query = query.filter_by(verification_status=status_filter)
    
    pagination = query.order_by(Guide.created_at.desc()).paginate(page=page, per_page=15, error_out=False)
    return render_template(
        "admin/guides.html", 
        guides=pagination.items, 
        pagination=pagination, 
        q=q, 
        status_filter=status_filter
    )


@admin_bp.route("/guides/new", methods=["GET", "POST"])
@login_required
@admin_required
def guide_create():
    # Get all users with 'Guide' role or allow selection
    users = User.query.filter(
        User.role.in_(['Guide', 'Admin', 'Super Admin']), 
        User.status == 'active'
    ).all()
    
    if request.method == "POST":
        user_id = request.form.get("user_id", type=int)
        if not user_id:
            flash("Please select a user for this guide.", "danger")
            return render_template("admin/guide_form.html", guide=None, users=users, title="Add Guide")
        
        # Check if user already has a guide profile
        existing = Guide.query.filter_by(user_id=user_id).first()
        if existing:
            flash("This user already has a guide profile.", "danger")
            return render_template("admin/guide_form.html", guide=None, users=users, title="Add Guide")
        
        try:
            g = Guide(
                user_id=user_id,
                phone=request.form.get("phone", "").strip(),  # <-- Added phone field
                bio=request.form.get("bio", ""),
                languages=request.form.get("languages", ""),
                experience=int(request.form.get("experience", 0)),
                hourly_rate=float(request.form.get("hourly_rate", 0)),
                daily_rate=float(request.form.get("daily_rate", 0)),
                specializations=request.form.get("specializations", ""),
                license_number=request.form.get("license_number", ""),
                verification_status=request.form.get("verification_status", "Pending"),
                status="active"
            )
            db.session.add(g)
            db.session.commit()
            log_action(f"Admin created guide for user #{user_id}", module="admin.guides", record_id=g.id)
            flash("Guide created successfully.", "success")
            return redirect(url_for("admin.guides"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error creating guide: {str(e)}", "danger")
    
    return render_template("admin/guide_form.html", guide=None, users=users, title="Add Guide")


@admin_bp.route("/guides/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def guide_edit(id):
    g = Guide.query.get_or_404(id)
    
    if request.method == "POST":
        try:
            g.bio = request.form.get("bio", g.bio)
            g.phone = request.form.get("phone", g.phone).strip()  # <-- Added phone field
            g.languages = request.form.get("languages", g.languages)
            g.experience = int(request.form.get("experience", g.experience))
            g.hourly_rate = float(request.form.get("hourly_rate", g.hourly_rate))
            g.daily_rate = float(request.form.get("daily_rate", g.daily_rate))
            g.specializations = request.form.get("specializations", g.specializations)
            g.license_number = request.form.get("license_number", g.license_number)
            g.verification_status = request.form.get("verification_status", g.verification_status)
            
            db.session.commit()
            log_action(f"Admin edited guide #{id}", module="admin.guides", record_id=id)
            flash("Guide updated successfully.", "success")
            return redirect(url_for("admin.guides"))
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating guide: {str(e)}", "danger")
    
    return render_template("admin/guide_form.html", guide=g, title="Edit Guide")


@admin_bp.route("/guides/<int:id>/verify", methods=["POST"])
@login_required
@admin_required
def guide_verify(id):
    g = Guide.query.get_or_404(id)
    action = request.form.get("action", "Approved")
    g.verification_status = action
    db.session.commit()
    log_action(f"Admin set guide #{id} status to {action}", module="admin.guides", record_id=id)
    flash(f"Guide {action}.", "success")
    return redirect(url_for("admin.guides"))


@admin_bp.route("/guides/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def guide_delete(id):
    g = Guide.query.get_or_404(id)
    g.soft_delete()
    db.session.commit()
    log_action(f"Admin deleted guide #{id}", module="admin.guides", record_id=id)
    flash("Guide removed.", "info")
    return redirect(url_for("admin.guides"))


# ─── HOTELS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/hotels")
@login_required
@admin_required
def hotels():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    query = Hotel.query.filter_by(status="active")
    if q:
        query = query.filter(Hotel.name.ilike(f"%{q}%"))
    pagination = query.paginate(page=page, per_page=15, error_out=False)
    return render_template("admin/hotels.html", hotels=pagination.items, pagination=pagination, q=q)


@admin_bp.route("/hotels/new", methods=["GET", "POST"])
@login_required
@admin_required
def hotel_create():
    if request.method == "POST":
        image = request.files.get("image")
        image_url = None
        if image and allowed_file(image.filename):
            image_url = save_uploaded_file(image, "hotels")
        h = Hotel(
            name=request.form.get("name", "").strip(),
            address=request.form.get("address", ""),
            description=request.form.get("description", ""),
            facilities=request.form.get("facilities", ""),
            price_per_night=float(request.form.get("price_per_night", 0)),
            image_url=image_url,
            latitude=float(request.form.get("latitude", 6.34)),
            longitude=float(request.form.get("longitude", 5.62)),
            contact=request.form.get("contact", ""),
            website=request.form.get("website", ""),
            stars=int(request.form.get("stars", 3)),
            total_rooms=int(request.form.get("total_rooms", 0)),
            featured=request.form.get("featured") == "on",
            status="active",
        )
        db.session.add(h)
        db.session.commit()
        log_action(f"Admin created hotel: {h.name}", module="admin.hotels", record_id=h.id)
        flash(f"Hotel '{h.name}' created.", "success")
        return redirect(url_for("admin.hotels"))
    return render_template("admin/hotel_form.html", hotel=None, title="New Hotel")


@admin_bp.route("/hotels/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def hotel_edit(id):
    h = Hotel.query.get_or_404(id)
    if request.method == "POST":
        h.name = request.form.get("name", h.name).strip()
        h.address = request.form.get("address", h.address)
        h.description = request.form.get("description", h.description)
        h.facilities = request.form.get("facilities", h.facilities)
        h.price_per_night = float(request.form.get("price_per_night", h.price_per_night))
        h.latitude = float(request.form.get("latitude", h.latitude))
        h.longitude = float(request.form.get("longitude", h.longitude))
        h.contact = request.form.get("contact", h.contact)
        h.website = request.form.get("website", h.website)
        h.stars = int(request.form.get("stars", h.stars))
        h.total_rooms = int(request.form.get("total_rooms", h.total_rooms))
        h.featured = request.form.get("featured") == "on"
        image = request.files.get("image")
        if image and allowed_file(image.filename):
            if h.image_url:
                delete_uploaded_file(h.image_url, "hotels")
            h.image_url = save_uploaded_file(image, "hotels")
        db.session.commit()
        log_action(f"Admin edited hotel #{id}", module="admin.hotels", record_id=id)
        flash("Hotel updated.", "success")
        return redirect(url_for("admin.hotels"))
    return render_template("admin/hotel_form.html", hotel=h, title="Edit Hotel")


@admin_bp.route("/hotels/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def hotel_delete(id):
    h = Hotel.query.get_or_404(id)
    if h.image_url:
        delete_uploaded_file(h.image_url, "hotels")
    h.soft_delete()
    db.session.commit()
    log_action(f"Admin deleted hotel #{id}", module="admin.hotels", record_id=id)
    flash("Hotel removed.", "info")
    return redirect(url_for("admin.hotels"))


# ─── RESTAURANTS ─────────────────────────────────────────────────────────────

@admin_bp.route("/restaurants")
@login_required
@admin_required
def restaurants():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    query = Restaurant.query.filter_by(status="active")
    if q:
        query = query.filter(Restaurant.name.ilike(f"%{q}%"))
    pagination = query.paginate(page=page, per_page=15, error_out=False)
    return render_template("admin/restaurants.html", restaurants=pagination.items, pagination=pagination, q=q)


@admin_bp.route("/restaurants/new", methods=["GET", "POST"])
@login_required
@admin_required
def restaurant_create():
    if request.method == "POST":
        image = request.files.get("image")
        image_url = None
        if image and allowed_file(image.filename):
            image_url = save_uploaded_file(image, "restaurants")
        r = Restaurant(
            name=request.form.get("name", "").strip(),
            address=request.form.get("address", ""),
            description=request.form.get("description", ""),
            cuisine=request.form.get("cuisine", ""),
            opening_hours=request.form.get("opening_hours", ""),
            image_url=image_url,
            latitude=float(request.form.get("latitude", 6.34)),
            longitude=float(request.form.get("longitude", 5.62)),
            contact=request.form.get("contact", ""),
            website=request.form.get("website", ""),
            price_range=request.form.get("price_range", "₦₦"),
            featured=request.form.get("featured") == "on",
            status="active",
        )
        db.session.add(r)
        db.session.commit()
        log_action(f"Admin created restaurant: {r.name}", module="admin.restaurants", record_id=r.id)
        flash(f"Restaurant '{r.name}' created.", "success")
        return redirect(url_for("admin.restaurants"))
    return render_template("admin/restaurant_form.html", restaurant=None, title="New Restaurant")


@admin_bp.route("/restaurants/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def restaurant_edit(id):
    r = Restaurant.query.get_or_404(id)
    if request.method == "POST":
        r.name = request.form.get("name", r.name).strip()
        r.address = request.form.get("address", r.address)
        r.description = request.form.get("description", r.description)
        r.cuisine = request.form.get("cuisine", r.cuisine)
        r.opening_hours = request.form.get("opening_hours", r.opening_hours)
        r.latitude = float(request.form.get("latitude", r.latitude))
        r.longitude = float(request.form.get("longitude", r.longitude))
        r.contact = request.form.get("contact", r.contact)
        r.price_range = request.form.get("price_range", r.price_range)
        r.featured = request.form.get("featured") == "on"
        image = request.files.get("image")
        if image and allowed_file(image.filename):
            if r.image_url:
                delete_uploaded_file(r.image_url, "restaurants")
            r.image_url = save_uploaded_file(image, "restaurants")
        db.session.commit()
        log_action(f"Admin edited restaurant #{id}", module="admin.restaurants", record_id=id)
        flash("Restaurant updated.", "success")
        return redirect(url_for("admin.restaurants"))
    return render_template("admin/restaurant_form.html", restaurant=r, title="Edit Restaurant")


@admin_bp.route("/restaurants/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def restaurant_delete(id):
    r = Restaurant.query.get_or_404(id)
    if r.image_url:
        delete_uploaded_file(r.image_url, "restaurants")
    r.soft_delete()
    db.session.commit()
    flash("Restaurant removed.", "info")
    return redirect(url_for("admin.restaurants"))


# ─── EVENTS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/events")
@login_required
@admin_required
def events():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    query = Event.query.filter_by(status="active")
    if q:
        query = query.filter(Event.name.ilike(f"%{q}%"))
    pagination = query.order_by(Event.date.desc()).paginate(page=page, per_page=15, error_out=False)
    return render_template("admin/events.html", events=pagination.items, pagination=pagination, q=q)


@admin_bp.route("/events/new", methods=["GET", "POST"])
@login_required
@admin_required
def event_create():
    if request.method == "POST":
        image = request.files.get("image")
        image_url = None
        if image and allowed_file(image.filename):
            image_url = save_uploaded_file(image, "events")
        date_str = request.form.get("date", "")
        end_str = request.form.get("end_date", "")
        try:
            date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M") if date_str else datetime.utcnow()
            end_date = datetime.strptime(end_str, "%Y-%m-%dT%H:%M") if end_str else None
        except ValueError:
            date = datetime.utcnow()
            end_date = None
        e = Event(
            name=request.form.get("name", "").strip(),
            description=request.form.get("description", ""),
            date=date, end_date=end_date,
            location=request.form.get("location", ""),
            image_url=image_url,
            ticket_price=float(request.form.get("ticket_price", 0)),
            capacity=int(request.form.get("capacity", 0)),
            organizer=request.form.get("organizer", ""),
            contact=request.form.get("contact", ""),
            featured=request.form.get("featured") == "on",
            active=True, status="active",
            category=request.form.get("category", ""),
        )
        db.session.add(e)
        db.session.commit()
        log_action(f"Admin created event: {e.name}", module="admin.events", record_id=e.id)
        flash(f"Event '{e.name}' created.", "success")
        return redirect(url_for("admin.events"))
    return render_template("admin/event_form.html", event=None, title="New Event")


@admin_bp.route("/events/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def event_edit(id):
    e = Event.query.get_or_404(id)
    if request.method == "POST":
        e.name = request.form.get("name", e.name).strip()
        e.description = request.form.get("description", e.description)
        e.location = request.form.get("location", e.location)
        e.ticket_price = float(request.form.get("ticket_price", e.ticket_price))
        e.capacity = int(request.form.get("capacity", e.capacity))
        e.organizer = request.form.get("organizer", e.organizer)
        e.contact = request.form.get("contact", e.contact)
        e.featured = request.form.get("featured") == "on"
        e.active = request.form.get("active") == "on"
        e.category = request.form.get("category", e.category)
        date_str = request.form.get("date", "")
        end_str = request.form.get("end_date", "")
        try:
            if date_str:
                e.date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M")
            if end_str:
                e.end_date = datetime.strptime(end_str, "%Y-%m-%dT%H:%M")
        except ValueError:
            pass
        image = request.files.get("image")
        if image and allowed_file(image.filename):
            if e.image_url:
                delete_uploaded_file(e.image_url, "events")
            e.image_url = save_uploaded_file(image, "events")
        db.session.commit()
        log_action(f"Admin edited event #{id}", module="admin.events", record_id=id)
        flash("Event updated.", "success")
        return redirect(url_for("admin.events"))
    return render_template("admin/event_form.html", event=e, title="Edit Event")


@admin_bp.route("/events/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def event_delete(id):
    e = Event.query.get_or_404(id)
    if e.image_url:
        delete_uploaded_file(e.image_url, "events")
    e.soft_delete()
    db.session.commit()
    flash("Event removed.", "info")
    return redirect(url_for("admin.events"))


# ─── NEWS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/news")
@login_required
@admin_required
def news():
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    query = News.query.filter_by(status="active")
    if q:
        query = query.filter(News.title.ilike(f"%{q}%"))
    pagination = query.order_by(News.created_at.desc()).paginate(page=page, per_page=15, error_out=False)
    return render_template("admin/news.html", articles=pagination.items, pagination=pagination, q=q)


@admin_bp.route("/news/new", methods=["GET", "POST"])
@login_required
@admin_required
def news_create():
    if request.method == "POST":
        image = request.files.get("image")
        image_url = None
        if image and allowed_file(image.filename):
            image_url = save_uploaded_file(image, "news")
        from ..utils.helpers import slugify
        import random
        title = request.form.get("title", "").strip()
        n = News(
            title=title,
            slug=slugify(title) + "-" + str(random.randint(1000, 9999)),
            content=request.form.get("content", ""),
            category=request.form.get("category", ""),
            tags=request.form.get("tags", ""),
            image_url=image_url,
            featured=request.form.get("featured") == "on",
            author=request.form.get("author", current_user.full_name),
            status="active",
        )
        db.session.add(n)
        db.session.commit()
        log_action(f"Admin created news: {n.title}", module="admin.news", record_id=n.id)
        flash("Article created.", "success")
        return redirect(url_for("admin.news"))
    return render_template("admin/news_form.html", article=None, title="New Article")


@admin_bp.route("/news/<int:id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def news_edit(id):
    n = News.query.get_or_404(id)
    if request.method == "POST":
        n.title = request.form.get("title", n.title).strip()
        n.content = request.form.get("content", n.content)
        n.category = request.form.get("category", n.category)
        n.tags = request.form.get("tags", n.tags)
        n.featured = request.form.get("featured") == "on"
        n.author = request.form.get("author", n.author)
        image = request.files.get("image")
        if image and allowed_file(image.filename):
            if n.image_url:
                delete_uploaded_file(n.image_url, "news")
            n.image_url = save_uploaded_file(image, "news")
        db.session.commit()
        log_action(f"Admin edited news #{id}", module="admin.news", record_id=id)
        flash("Article updated.", "success")
        return redirect(url_for("admin.news"))
    return render_template("admin/news_form.html", article=n, title="Edit Article")


@admin_bp.route("/news/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def news_delete(id):
    n = News.query.get_or_404(id)
    if n.image_url:
        delete_uploaded_file(n.image_url, "news")
    n.soft_delete()
    db.session.commit()
    flash("Article removed.", "info")
    return redirect(url_for("admin.news"))


# ─── BOOKINGS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/bookings")
@login_required
@admin_required
def bookings():
    page = request.args.get("page", 1, type=int)
    status_filter = request.args.get("status", "")
    btype = request.args.get("type", "")
    query = Booking.query.filter_by(status="active")
    if status_filter:
        query = query.filter_by(booking_status=status_filter)
    if btype:
        query = query.filter_by(booking_type=btype)
    pagination = query.order_by(Booking.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template("admin/bookings.html", bookings=pagination.items, pagination=pagination, status_filter=status_filter, btype=btype)


@admin_bp.route("/bookings/<int:id>/status", methods=["POST"])
@login_required
@admin_required
def booking_status(id):
    b = Booking.query.get_or_404(id)
    new_status = request.form.get("status", "")
    allowed = ["Pending", "Confirmed", "Completed", "Cancelled"]
    if new_status in allowed:
        b.booking_status = new_status
        db.session.commit()
        log_action(f"Admin changed booking #{id} status to {new_status}", module="admin.bookings", record_id=id)
        flash(f"Booking status updated to {new_status}.", "success")
    return redirect(url_for("admin.bookings"))


# ─── REVIEWS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/reviews")
@login_required
@admin_required
def reviews():
    page = request.args.get("page", 1, type=int)
    query = Review.query.filter_by(status="active").order_by(Review.created_at.desc())
    pagination = query.paginate(page=page, per_page=20, error_out=False)
    return render_template("admin/reviews.html", reviews=pagination.items, pagination=pagination)


@admin_bp.route("/reviews/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def review_delete(id):
    r = Review.query.get_or_404(id)
    r.soft_delete()
    db.session.commit()
    flash("Review removed.", "info")
    return redirect(url_for("admin.reviews"))


@admin_bp.route("/reviews/<int:id>/approve", methods=["POST"])
@login_required
@admin_required
def review_approve(id):
    r = Review.query.get_or_404(id)
    r.is_approved = not r.is_approved
    db.session.commit()
    flash("Review approval toggled.", "success")
    return redirect(url_for("admin.reviews"))


# ─── ANALYTICS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/analytics")
@login_required
@admin_required
def analytics():
    # Monthly bookings (last 12 months)
    monthly = []
    for i in range(11, -1, -1):
        month_start = (datetime.utcnow().replace(day=1) - timedelta(days=i * 30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1)
        count = Booking.query.filter(Booking.created_at >= month_start, Booking.created_at < month_end).count()
        revenue = db.session.query(db.func.sum(Booking.total_price)).filter(
            Booking.created_at >= month_start, Booking.created_at < month_end,
            Booking.booking_status.in_(["Confirmed", "Completed"])
        ).scalar() or 0
        monthly.append({"month": month_start.strftime("%b %Y"), "count": count, "revenue": float(revenue)})

    top_attractions = Attraction.query.filter_by(status="active").order_by(Attraction.views.desc()).limit(10).all()
    top_guides = Guide.query.filter_by(verification_status="Approved", status="active").order_by(Guide.total_tours.desc()).limit(5).all()
    qr_stats = QRCode.query.order_by(QRCode.scanned_count.desc()).limit(10).all()

    booking_types = db.session.query(Booking.booking_type, db.func.count(Booking.id)).group_by(Booking.booking_type).all()
    booking_statuses = db.session.query(Booking.booking_status, db.func.count(Booking.id)).group_by(Booking.booking_status).all()

    role_counts = db.session.query(User.role, db.func.count(User.id)).filter(User.status != "deleted").group_by(User.role).all()

    return render_template(
        "admin/analytics.html",
        monthly=json.dumps(monthly),
        top_attractions=top_attractions,
        top_guides=top_guides,
        qr_stats=qr_stats,
        booking_types=json.dumps([{"type": bt, "count": c} for bt, c in booking_types]),
        booking_statuses=json.dumps([{"status": bs, "count": c} for bs, c in booking_statuses]),
        role_counts=json.dumps([{"role": r, "count": c} for r, c in role_counts]),
        total_revenue=sum(m["revenue"] for m in monthly),
        total_bookings=sum(m["count"] for m in monthly),
    )


# ─── ANALYTICS EXPORT ──────────────────────────────────────────────────────

@admin_bp.route("/analytics/export/csv")
@login_required
@admin_required
def export_csv():
    bookings = Booking.query.filter_by(status="active").order_by(Booking.created_at.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Reference", "User", "Type", "Target", "Start Date", "Total Price", "Status", "Created"])
    for b in bookings:
        writer.writerow([b.id, b.reference_code, b.user.full_name if b.user else "", b.booking_type, b.target_name, b.start_date.strftime("%Y-%m-%d"), f"₦{b.total_price:,.0f}", b.booking_status, b.created_at.strftime("%Y-%m-%d")])
    output.seek(0)
    log_action("Admin exported bookings CSV", module="admin.analytics")
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=edo_odyssey_bookings.csv"})


@admin_bp.route("/analytics/export/excel")
@login_required
@admin_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a3a6b", end_color="1a3a6b", fill_type="solid")
        headers = ["ID", "Reference", "User", "Type", "Target", "Start Date", "Price (₦)", "Status", "Date Created"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        bookings = Booking.query.filter_by(status="active").all()
        for row, b in enumerate(bookings, 2):
            ws.cell(row=row, column=1, value=b.id)
            ws.cell(row=row, column=2, value=b.reference_code)
            ws.cell(row=row, column=3, value=b.user.full_name if b.user else "")
            ws.cell(row=row, column=4, value=b.booking_type)
            ws.cell(row=row, column=5, value=b.target_name)
            ws.cell(row=row, column=6, value=b.start_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=7, value=b.total_price)
            ws.cell(row=row, column=8, value=b.booking_status)
            ws.cell(row=row, column=9, value=b.created_at.strftime("%Y-%m-%d"))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        log_action("Admin exported bookings Excel", module="admin.analytics")
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="edo_odyssey_bookings.xlsx")
    except ImportError:
        flash("openpyxl not installed.", "danger")
        return redirect(url_for("admin.analytics"))


# ─── CMS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/cms")
@login_required
@admin_required
def cms():
    settings = CMSSetting.query.order_by(CMSSetting.group, CMSSetting.key).all()
    faqs = FAQ.query.order_by(FAQ.order).all()
    partners = Partner.query.order_by(Partner.type, Partner.name).all()
    return render_template("admin/cms.html", settings=settings, faqs=faqs, partners=partners)


@admin_bp.route("/cms/update", methods=["POST"])
@login_required
@admin_required
def cms_update():
    for key, value in request.form.items():
        if key.startswith("_"):
            continue
        s = CMSSetting.query.filter_by(key=key).first()
        if s:
            s.value = value
        else:
            db.session.add(CMSSetting(key=key, value=value))
    db.session.commit()
    log_action("Admin updated CMS settings", module="admin.cms")
    flash("CMS settings updated across all pages.", "success")
    return redirect(url_for("admin.cms"))


@admin_bp.route("/cms/faq/new", methods=["POST"])
@login_required
@admin_required
def faq_create():
    f = FAQ(
        question=request.form.get("question", ""),
        answer=request.form.get("answer", ""),
        category=request.form.get("category", "general"),
        order=int(request.form.get("order", 0)),
        active=True,
    )
    db.session.add(f)
    db.session.commit()
    flash("FAQ added.", "success")
    return redirect(url_for("admin.cms"))


@admin_bp.route("/cms/faq/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def faq_delete(id):
    f = FAQ.query.get_or_404(id)
    db.session.delete(f)
    db.session.commit()
    flash("FAQ deleted.", "info")
    return redirect(url_for("admin.cms"))


@admin_bp.route("/cms/partner/new", methods=["POST"])
@login_required
@admin_required
def partner_create():
    p = Partner(
        name=request.form.get("name", ""),
        logo_url=request.form.get("logo_url", ""),
        website=request.form.get("website", ""),
        description=request.form.get("description", ""),
        type=request.form.get("type", "partner"),
        active=True,
    )
    db.session.add(p)
    db.session.commit()
    flash("Partner/Sponsor added.", "success")
    return redirect(url_for("admin.cms"))


@admin_bp.route("/cms/partner/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def partner_delete(id):
    p = Partner.query.get_or_404(id)
    db.session.delete(p)
    db.session.commit()
    flash("Partner removed.", "info")
    return redirect(url_for("admin.cms"))


# ─── AUDIT LOGS ─────────────────────────────────────────────────────────────────

@admin_bp.route("/audit-logs")
@login_required
@admin_required
def audit_logs():
    page = request.args.get("page", 1, type=int)
    query = AuditLog.query.order_by(AuditLog.created_at.desc())
    pagination = query.paginate(page=page, per_page=30, error_out=False)
    return render_template("admin/audit_logs.html", logs=pagination.items, pagination=pagination)


# ─── CATEGORIES ─────────────────────────────────────────────────────────────────

@admin_bp.route("/categories")
@login_required
@admin_required
def categories():
    cats = Category.query.filter_by(status="active").all()
    return render_template("admin/categories.html", categories=cats)


@admin_bp.route("/categories/new", methods=["POST"])
@login_required
@admin_required
def category_create():
    c = Category(
        name=request.form.get("name", "").strip(),
        description=request.form.get("description", ""),
        type=request.form.get("type", "attraction"),
        icon=request.form.get("icon", "bi-tag"),
        color=request.form.get("color", "#1a3a6b"),
        status="active",
    )
    db.session.add(c)
    db.session.commit()
    flash("Category created.", "success")
    return redirect(url_for("admin.categories"))


@admin_bp.route("/categories/<int:id>/delete", methods=["POST"])
@login_required
@admin_required
def category_delete(id):
    c = Category.query.get_or_404(id)
    c.soft_delete()
    db.session.commit()
    flash("Category removed.", "info")
    return redirect(url_for("admin.categories"))


# ─── DOWNLOAD PROJECT ZIP ─────────────────────────────────────────────────────

@admin_bp.route("/download-zip")
@login_required
@admin_required
def download_zip():
    """Create and return a ZIP of the entire project source code."""
    project_root = os.path.abspath(os.path.join(current_app.root_path, ".."))
    buf = io.BytesIO()
    exclude_dirs = {"__pycache__", ".git", "venv", "env", ".env", "node_modules", "uploads"}
    exclude_exts = {".pyc", ".pyo", ".db", ".sqlite", ".egg-info"}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(project_root):
            dirs[:] = [d for d in dirs if d not in exclude_dirs]
            for file in files:
                if any(file.endswith(ext) for ext in exclude_exts):
                    continue
                full_path = os.path.join(root, file)
                arcname = os.path.relpath(full_path, project_root)
                zf.write(full_path, arcname)
    buf.seek(0)
    log_action("Admin downloaded project ZIP", module="admin.download")
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name="edo_odyssey_source.zip",
    )


# ─── NOTIFICATIONS BROADCAST ──────────────────────────────────────────────────

@admin_bp.route("/notify-all", methods=["POST"])
@login_required
@admin_required
def notify_all():
    title = request.form.get("title", "")
    message = request.form.get("message", "")
    if title and message:
        users = User.query.filter(User.status == "active").all()
        for u in users:
            n = Notification(user_id=u.id, title=title, message=message, type="info")
            db.session.add(n)
        db.session.commit()
        log_action(f"Admin broadcast notification to all users: {title}", module="admin.notifications")
        flash(f"Notification sent to {len(users)} users.", "success")
    return redirect(url_for("admin.dashboard"))